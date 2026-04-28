# app/api/v1/routers/persons.py
"""
ИСПРАВЛЕНИЕ: Вся таблица persons в RAM при импорте.

Проблема (старый код):
    existing_records = {p.full_name.lower(): p for p in db.query(Person).all()}
    При 10 000+ записях — вся таблица грузилась в память Python-процесса.

Решение:
    Используем PostgreSQL INSERT ... ON CONFLICT DO UPDATE (upsert) через
    sqlalchemy.dialects.postgresql.insert.
    БД сама обрабатывает конфликты по full_name — Python не держит таблицу в памяти.

    Логика при совпадении full_name:
      - rank, doc_number, department и прочие поля обновляются только если
        в БД значение NULL (COALESCE — не затираем заполненные данные).
      - updated_at обновляется всегда.

    Результат: O(1) память вместо O(N), один SQL-запрос на батч вместо N+1.

    Также исправлен upsert_person_from_slot — теперь тоже использует pg_insert
    вместо SELECT + условного INSERT.
"""

import io
from datetime import date, datetime, timezone
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Query, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from pydantic import BaseModel, Field, ConfigDict

from app.db.database import get_db
from app.models.person import Person
from app.models.user import User
from app.models.duty import DutySchedulePerson
from app.api.dependencies import get_current_user, get_current_active_admin, require_permission
from app.core.audit import notify_all_admins
from app.core.websockets import manager

router = APIRouter()

# ─── Константы импорта ────────────────────────────────────────────────────────

CHUNK_SIZE  = 500
MAX_FILE_MB = 10
MAX_ROWS    = 10_000

TEMPLATE_COLS = [
    ("ФИО",             "full_name",      True,  "Иванов Иван Иванович"),
    ("Воинское звание", "rank",           True,  "Майор"),
    ("Номер документа", "doc_number",     False,  "АА 123456"),
    ("Подразделение",   "department",     False, "управление_1"),
    ("Должность",       "position_title", False, "Начальник отдела"),
    ("Дата рождения",   "birth_date",     False, "01.01.1985"),
    ("Телефон",         "phone",          False, "+7 (999) 123-45-67"),
    ("Примечание",      "notes",          False, "Любая заметка"),
]

COL_EXAMPLE = [c[3] for c in TEMPLATE_COLS]


# ─── Схемы ────────────────────────────────────────────────────────────────────

class PersonCreate(BaseModel):
    full_name:      str            = Field(...,  min_length=2, max_length=300, strip_whitespace=True)
    rank:           Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    doc_number:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    department:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    position_title: Optional[str]  = Field(None, max_length=200, strip_whitespace=True)
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = Field(None, max_length=50,  strip_whitespace=True)
    notes:          Optional[str]  = Field(None, max_length=2000, strip_whitespace=True)


class PersonUpdate(BaseModel):
    full_name:      Optional[str]  = Field(None, min_length=2, max_length=300, strip_whitespace=True)
    rank:           Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    doc_number:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    department:     Optional[str]  = Field(None, max_length=100, strip_whitespace=True)
    position_title: Optional[str]  = Field(None, max_length=200, strip_whitespace=True)
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = Field(None, max_length=50,  strip_whitespace=True)
    notes:          Optional[str]  = Field(None, max_length=2000, strip_whitespace=True)


class PersonResponse(BaseModel):
    id:             int
    full_name:      str
    rank:           Optional[str]  = None
    doc_number:     Optional[str]  = None
    department:     Optional[str]  = None
    position_title: Optional[str]  = None
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = None
    notes:          Optional[str]  = None
    fired_at:       Optional[datetime] = None   # NULL → активный

    model_config = ConfigDict(from_attributes=True)


class PersonSuggestion(BaseModel):
    """
    Предложение совпадения из общей базы людей.

    match_score: 0-100, насколько сильно совпадает с тем что вводит пользователь.
      100 → точное совпадение full_name + rank + doc_number
      >=85 → очень высокая уверенность, UI может авто-подставлять
      >=60 → похожее ФИО, но детали различаются — показать как опцию
      <60  → скорее шум, фильтруется на бэке

    is_exact: true если full_name совпадает побуквенно (регистронезависимо).
      Фронтенд использует это чтобы подчеркнуть точное совпадение отдельно
      от fuzzy-вариантов.

    birth_date / phone — возвращаются чтобы единый fio_autocomplete.js мог
    при выборе подставить сразу все поля (форма добавления в dept_persons,
    редактор слотов и т.д.), а не делать второй запрос за деталями.
    """
    id:             int
    full_name:      str
    rank:           Optional[str]  = None
    doc_number:     Optional[str]  = None
    department:     Optional[str]  = None
    position_title: Optional[str]  = None
    birth_date:     Optional[date] = None
    phone:          Optional[str]  = None
    notes:          Optional[str]  = None
    match_score:    int
    is_exact:       bool

    model_config = ConfigDict(from_attributes=True)


class ImportRowError(BaseModel):
    row:     int
    message: str


class ImportResult(BaseModel):
    message:  str
    added:    int
    updated:  int
    skipped:  int
    errors:   List[ImportRowError] = []


# ─── Вспомогательные функции парсинга ────────────────────────────────────────

def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() != "none" else None


def _parse_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _validate_row(row_num: int, fields: dict) -> Optional[str]:
    fn = fields.get("full_name")
    if not fn or len(fn) < 2:
        return f"Строка {row_num}: ФИО обязательно (минимум 2 символа)"
    if not fields.get("rank"):
        return f"Строка {row_num}: Воинское звание обязательно (ФИО: {fn})"
    return None


# ─── Поиск (автодополнение) ───────────────────────────────────────────────────

@router.get(
    "/search",
    response_model=List[PersonResponse],
    summary="Поиск человека по ФИО (для автодополнения)",
)
def search_persons(
        q:     str = Query(..., min_length=2),
        limit: int = Query(10, ge=1, le=50),
        db:    Session = Depends(get_db),
        current_user: User = Depends(get_current_user),
):
    """
    Поиск по ФИО (автодополнение при заполнении слотов).

    Админ видит всех.
    Управление видит:
      - своих людей (department == username)
      - «общих» людей без привязки к управлению (department IS NULL)
        — это например те, кого другие департаменты ещё не пометили,
           либо которых админ импортировал без указания управления.

    Уволенные (fired_at IS NOT NULL) не попадают в выдачу — ни для admin,
    ни для управлений. Для «показать уволенных» используйте
    GET /persons?include_fired=true.
    """
    from sqlalchemy import or_

    query = (
        db.query(Person)
          .filter(Person.full_name.ilike(f"%{q}%"))
          .filter(Person.fired_at.is_(None))
    )
    if current_user.role != "admin":
        query = query.filter(
            or_(
                Person.department == current_user.username,
                Person.department.is_(None),
            )
        )
    return query.order_by(Person.full_name).limit(limit).all()


# ─── Подбор ФИО: поиск совпадений в ОБЩЕЙ базе ───────────────────────────────

@router.get(
    "/suggest",
    response_model=List[PersonSuggestion],
    summary="Подбор человека из общей базы по ФИО (fuzzy)",
)
def suggest_persons(
        full_name:  str           = Query(..., min_length=2, max_length=300),
        rank:       Optional[str] = Query(None, max_length=100),
        doc_number: Optional[str] = Query(None, max_length=100),
        limit:      int           = Query(5, ge=1, le=20),
        db:         Session       = Depends(get_db),
        current_user: User        = Depends(get_current_user),
):
    """
    Возвращает кандидатов из ОБЩЕЙ базы (таблица persons).

    Ключевой сценарий:
      Пользователь добавляет человека в свою таблицу (заполняет слот).
      Фронт дёргает /persons/suggest с тем что ввели.
      Если в общей базе уже есть такой человек — возвращаем предложение
      с match_score. UI может:
        • при score >= 85 — автоподставить все поля и показать плашку
          «найден в общей базе управления X»;
        • при 60-84 — показать список кандидатов для ручного выбора;
        • при <60 — не мешать, пусть вводит новое.
      Если никто не подошёл — эндпоинт вернёт пустой список, и фронт
      сохранит ввод как новую запись через существующий upsert.

    Это ВАЖНО: сама логика создания/обновления persons не меняется.
    Endpoint /suggest — чисто read-only, подсказка.

    Алгоритм score:
      • База — pg_trgm similarity(lower(full_name), lower(:q)) × 100.
        Trigram устойчив к опечаткам, перестановке слов (Иванов Иван vs
        Иван Иванов), разнице пробелов.
      • +10 бонус если совпал rank (точно, без регистра).
      • +15 бонус если совпал doc_number (doc_number — сильный сигнал).
      • Потолок 100.
      • Точное совпадение по lower(full_name) всегда даёт 100 и is_exact.

    Видимость:
      • admin: видит всех кандидатов.
      • department: видит только своих (department == username) и «общих»
        (department IS NULL). Чужих управлений не видит — переводы между
        управлениями централизованы у админа, поэтому department'у не
        нужно знать что человек уже закреплён за другим управлением.
        При попытке создать дубликат POST /persons вернёт 409.
    """
    q = full_name.strip()
    if len(q) < 2:
        return []

    # SQL: similarity() из pg_trgm даёт число 0..1.
    # Коалесцируем rank/doc_number на пустую строку, иначе NULL != NULL
    # и бонус не начислится.
    #
    # Параметры передаются bind-параметрами (:name_q, :rank_q, :doc_q) —
    # никакой конкатенации с пользовательскими данными → SQL-инъекция
    # невозможна.
    # Score-логика: trigram-similarity плохо работает для коротких префиксов
    # ("ярощ" против "Ярощук Александр Павлович" даёт ~15%, хотя совпадение
    # очевидное). Поэтому если запрос — подстрока full_name, даём ему пол
    # 65 (достаточно чтобы пройти порог), а префикс слова — ещё +10 сверху.
    # Итоговый score = MAX(trigram×100 + бонусы, substring-score + бонусы).
    sql = """
        SELECT
            id, full_name, rank, doc_number, department, position_title,
            birth_date, phone, notes,
            GREATEST(
                LEAST(
                    GREATEST(
                        ROUND(similarity(lower(full_name), lower(:name_q)) * 100)::int,
                        CASE
                            WHEN lower(full_name) LIKE lower(:name_q) || '%'
                              OR lower(full_name) LIKE '% ' || lower(:name_q) || '%'
                              THEN 75
                            WHEN lower(full_name) LIKE '%' || lower(:name_q) || '%'
                              THEN 60
                            ELSE 0
                        END
                    )
                    + CASE
                          WHEN :rank_q <> '' AND lower(COALESCE(rank, '')) = lower(:rank_q)
                          THEN 10 ELSE 0
                      END
                    + CASE
                          WHEN :doc_q <> '' AND COALESCE(doc_number, '') = :doc_q
                          THEN 15 ELSE 0
                      END,
                    100
                ),
                0
            ) AS score,
            (lower(full_name) = lower(:name_q)) AS is_exact
        FROM persons
        WHERE
            fired_at IS NULL
            AND (
                similarity(lower(full_name), lower(:name_q)) > 0.25
                OR lower(full_name) LIKE '%' || lower(:name_q) || '%'
            )
    """
    params: dict = {
        "name_q": q,
        "rank_q": (rank or "").strip(),
        "doc_q":  (doc_number or "").strip(),
    }

    # Видимость для не-админа: только свои + общие.
    # Чужих не показываем — переводы централизованы у админа.
    # "Общий" = department IS NULL ИЛИ пустая строка (исторически
    # в БД встречаются оба варианта — импорт Excel кладёт '', ручное
    # создание NULL). COALESCE нормализует обоих.
    if current_user.role != "admin":
        sql += (
            " AND (department = :user_dept OR COALESCE(department, '') = '')"
        )
        params["user_dept"] = current_user.username

    sql += """
        ORDER BY
            is_exact DESC,
            score DESC,
            full_name ASC
        LIMIT :lim
    """
    params["lim"] = limit

    rows = db.execute(text(sql), params).mappings().all()

    # Отсекаем слабых кандидатов (score < 40). Они попадают в выборку
    # за счёт LIKE-подстроки, но для UI это шум.
    result: List[PersonSuggestion] = []
    for r in rows:
        if r["score"] < 40 and not r["is_exact"]:
            continue
        result.append(PersonSuggestion(
            id=             r["id"],
            full_name=      r["full_name"],
            rank=           r["rank"],
            doc_number=     r["doc_number"],
            department=     r["department"],
            position_title= r["position_title"],
            birth_date=     r["birth_date"],
            phone=          r["phone"],
            notes=          r["notes"],
            match_score=    int(r["score"]),
            is_exact=       bool(r["is_exact"]),
        ))

    return result


# ─── Скачать шаблон Excel ─────────────────────────────────────────────────────

@router.get(
    "/import/template",
    summary="Скачать шаблон Excel для импорта",
)
def download_import_template(
        current_user: User = Depends(get_current_active_admin),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Шаблон импорта"

    RED_FILL    = PatternFill("solid", fgColor="C0392B")
    GREY_FILL   = PatternFill("solid", fgColor="7F8C8D")
    YELLOW_FILL = PatternFill("solid", fgColor="F9E79F")
    THIN        = Side(style="thin", color="CCCCCC")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    ws.merge_cells("A1:H1")
    note = ws["A1"]
    note.value = (
        "🔴 Красные колонки — ОБЯЗАТЕЛЬНЫЕ.  "
        "⚫ Серые — необязательные.  "
        "Строку 2 (пример) можно удалить.  "
        "Данные начинаются со строки 3."
    )
    note.font      = Font(bold=True, size=10)
    note.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for col_idx, (header, _, required, _ex) in enumerate(TEMPLATE_COLS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill      = RED_FILL if required else GREY_FILL
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[2].height = 22

    for col_idx, (_h, _f, _r, example) in enumerate(TEMPLATE_COLS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=example)
        cell.fill      = YELLOW_FILL
        cell.font      = Font(italic=True, color="555555", size=10)
        cell.alignment = Alignment(vertical="center")
        cell.border    = BORDER
    ws.row_dimensions[3].height = 18

    for i, width in enumerate([35, 20, 16, 18, 28, 16, 22, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="persons_template.xlsx"'},
    )


# ─── Массовый импорт из Excel ─────────────────────────────────────────────────

@router.post(
    "/import",
    response_model=ImportResult,
    summary="Массовый импорт из Excel",
)
async def import_persons_from_excel(
        file: UploadFile = File(...),
        db:   Session    = Depends(get_db),
        current_admin: User = Depends(get_current_active_admin),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Формат файла должен быть .xlsx")

    contents = await file.read()

    if len(contents) > MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(
            status_code=400,
            detail=f"Файл слишком большой (максимум {MAX_FILE_MB} МБ)",
        )

    try:
        wb    = openpyxl.load_workbook(io.BytesIO(contents), data_only=True, read_only=True)
        sheet = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось прочитать Excel файл")

    # ── Парсинг и валидация строк ─────────────────────────────────────────────
    errors:       List[ImportRowError] = []
    valid_rows:   List[dict]           = []
    seen_names:   set                  = set()
    skipped_count = 0
    row_num       = 0

    for excel_row in sheet.iter_rows(min_row=2, values_only=True):
        row_num += 1

        if row_num > MAX_ROWS:
            errors.append(ImportRowError(
                row=row_num,
                message=f"Превышен лимит {MAX_ROWS} строк. Остаток проигнорирован."
            ))
            break

        if not any(excel_row):
            skipped_count += 1
            continue

        def get_col(idx: int) -> Optional[str]:
            return _clean(excel_row[idx]) if len(excel_row) > idx else None

        first_val = get_col(0)

        if not first_val:
            skipped_count += 1
            continue

        # Пропускаем строку-заголовок и строку-пример шаблона
        if first_val.lower() in ("фио", "fullname", "full_name") or first_val == COL_EXAMPLE[0]:
            skipped_count += 1
            continue

        fields = {
            "full_name":      first_val,
            "rank":           get_col(1),
            "doc_number":     get_col(2),
            "department":     get_col(3),
            "position_title": get_col(4),
            "birth_date":     _parse_date(excel_row[5]) if len(excel_row) > 5 else None,
            "phone":          get_col(6),
            "notes":          get_col(7),
        }

        err_msg = _validate_row(row_num + 1, fields)
        if err_msg:
            errors.append(ImportRowError(row=row_num + 1, message=err_msg))
            continue

        # Дедупликация внутри самого файла
        key = fields["full_name"].lower()
        if key in seen_names:
            errors.append(ImportRowError(
                row=row_num + 1,
                message=f"Дубль ФИО «{fields['full_name']}» внутри файла — пропущена"
            ))
            skipped_count += 1
            continue
        seen_names.add(key)

        now = datetime.now(timezone.utc)
        valid_rows.append({
            "full_name":      fields["full_name"],
            "rank":           fields["rank"],
            "doc_number":     fields["doc_number"],
            "department":     fields["department"],
            "position_title": fields["position_title"],
            "birth_date":     fields["birth_date"],
            "phone":          fields["phone"],
            "notes":          fields["notes"],
            "created_at":     now,
            "updated_at":     now,
        })

    if not valid_rows:
        return ImportResult(
            message="Нет валидных строк для импорта.",
            added=0, updated=0, skipped=skipped_count, errors=errors,
        )

    # ── PostgreSQL upsert батчами ─────────────────────────────────────────────
    # ИСПРАВЛЕНО: один SQL-запрос на батч вместо загрузки всей таблицы в RAM.
    #
    # INSERT ... ON CONFLICT (full_name) DO UPDATE SET
    #   rank = COALESCE(persons.rank, EXCLUDED.rank),  ← не затираем если уже есть
    #   ...
    #
    # xmax = 0 означает что строка была вставлена (INSERT),
    # xmax != 0 — что обновлена (UPDATE). Так считаем added/updated без SELECT.

    added_count   = 0
    updated_count = 0

    for i in range(0, len(valid_rows), CHUNK_SIZE):
        chunk = valid_rows[i: i + CHUNK_SIZE]

        stmt = pg_insert(Person).values(chunk)
        stmt = stmt.on_conflict_do_update(
            index_elements=["full_name"],
            set_={
                # COALESCE: берём новое значение только если старое NULL
                "rank":           text("COALESCE(persons.rank,           EXCLUDED.rank)"),
                "doc_number":     text("COALESCE(persons.doc_number,     EXCLUDED.doc_number)"),
                "department":     text("COALESCE(persons.department,     EXCLUDED.department)"),
                "position_title": text("COALESCE(persons.position_title, EXCLUDED.position_title)"),
                "birth_date":     text("COALESCE(persons.birth_date,     EXCLUDED.birth_date)"),
                "phone":          text("COALESCE(persons.phone,          EXCLUDED.phone)"),
                "notes":          text("COALESCE(persons.notes,          EXCLUDED.notes)"),
                "updated_at":     stmt.excluded.updated_at,
            },
        )
        # returning xmax: 0 = INSERT, !=0 = UPDATE
        stmt = stmt.returning(
            Person.id,
            text("(xmax = 0)::int AS was_inserted"),
        )

        try:
            result = db.execute(stmt)
            db.commit()
            for row in result.fetchall():
                if row[1]:
                    added_count += 1
                else:
                    updated_count += 1
        except Exception as e:
            db.rollback()
            errors.append(ImportRowError(
                row=0,
                message=f"Ошибка батча {i+1}–{i+len(chunk)}: {str(e)[:120]}"
            ))

    return ImportResult(
        message=(
            f"Импорт завершён. "
            f"Добавлено: {added_count}, обновлено: {updated_count}, "
            f"пропущено: {skipped_count}, ошибок: {len(errors)}."
        ),
        added=added_count,
        updated=updated_count,
        skipped=skipped_count,
        errors=errors,
    )


# ─── CRUD ─────────────────────────────────────────────────────────────────────

@router.get("", summary="Получить базу людей")
def get_all_persons(
        db:   Session = Depends(get_db),
        current_user: User = Depends(require_permission("persons")),
        skip:  int = Query(0, ge=0),
        limit: int = Query(500, ge=1, le=2000),
        q:     Optional[str] = Query(None),
        page:  Optional[int] = Query(None, ge=1),
        sort:  str = Query("full_name", regex="^(full_name|rank|doc_number|department|created_at)$"),
        order: str = Query("asc", regex="^(asc|desc)$"),
        unassigned:    bool = Query(False),
        mine:          bool = Query(False),
        include_fired: bool = Query(False, description="Показать уволенных (admin-only UX)"),
):
    """
    База людей с гибким режимом ответа.

    Три режима видимости для department-пользователя:
      • mine=true        → только свои (department == username)      → "Мои люди"
      • unassigned=true  → только общие (department IS NULL)          → "Личный состав"
      • оба false        → свои + общие (объединённый режим по умолч.) → автодополнение/поиск

    Для admin:
      • unassigned=true → только общие
      • иначе           → все записи

    Формат ответа:
      • Без ?page=   — плоский List[PersonResponse] (старый контракт).
      • С ?page=N    — {items, total, page, pages, limit} для dept_persons.js.

    Permission: требуется "persons" в user.permissions (admin всегда ок).

    ИСПРАВЛЕНО: раньше dept_persons.js в режиме 'mine' не передавал
    никакой флаг, и бэк отдавал "свои + общие" — из-за этого во вкладке
    «Мои люди» показывались и чужие общие записи. Теперь mine=true даёт
    строгую фильтрацию только по своему управлению.
    """
    from sqlalchemy import or_, desc as sa_desc, asc as sa_asc

    # Взаимоисключающие флаги: если оба true — ошибка логики фронта
    if mine and unassigned:
        raise HTTPException(
            status_code=400,
            detail="Нельзя одновременно указывать mine=true и unassigned=true",
        )

    query = db.query(Person)

    # По умолчанию скрываем уволенных. Admin может запросить их отдельно
    # через include_fired=true (для вкладки «Уволенные»); department не
    # должен видеть уволенных ни в каком режиме — это централизованно
    # управляемое поле.
    if not include_fired or current_user.role != "admin":
        query = query.filter(Person.fired_at.is_(None))

    # Видимость по роли
    if current_user.role != "admin":
        if mine:
            # Строгий режим: только свои (department в точности == username).
            # Даже admin здесь видит только то что сам бы создал от своего
            # имени, но admin до этой ветки не доходит.
            query = query.filter(Person.department == current_user.username)
        elif unassigned:
            query = query.filter(Person.department.is_(None))
        else:
            query = query.filter(
                or_(
                    Person.department == current_user.username,
                    Person.department.is_(None),
                )
            )
    else:
        # admin: mine бессмысленно (у него нет своей квоты), unassigned работает
        if unassigned:
            query = query.filter(Person.department.is_(None))

    if q:
        query = query.filter(Person.full_name.ilike(f"%{q}%"))

    # Сортировка
    sort_col = getattr(Person, sort, Person.full_name)
    direction = sa_desc if order == "desc" else sa_asc
    query = query.order_by(direction(sort_col).nullslast(), Person.id.asc())

    # Paginated режим
    if page is not None:
        total = query.count()
        pages = max(1, (total + limit - 1) // limit)
        items = query.offset((page - 1) * limit).limit(limit).all()
        return {
            "items":  [PersonResponse.model_validate(p) for p in items],
            "total":  total,
            "page":   page,
            "pages":  pages,
            "limit":  limit,
        }

    # Flat list режим (backward compatible)
    rows = query.offset(skip).limit(limit).all()
    return [PersonResponse.model_validate(p) for p in rows]


@router.post("", response_model=PersonResponse, status_code=status.HTTP_201_CREATED,
             summary="Добавить человека")
async def create_person(
        person_in: PersonCreate,
        db:        Session = Depends(get_db),
        current_user: User = Depends(require_permission("persons")),
):
    department = person_in.department if current_user.role == "admin" else current_user.username

    existing = db.query(Person).filter(Person.full_name.ilike(person_in.full_name)).first()
    if existing:
        # Осиротевшего (никому не принадлежит, не уволен) подхватываем текущему
        # пользователю — типичный кейс: человек попал в БД через слот/импорт
        # без department, и теперь управление «забирает» его к себе.
        if existing.department is None and existing.fired_at is None:
            existing.department = department
            db.commit()
            db.refresh(existing)
            return existing
        owner  = existing.department or "—"
        suffix = " (уволен)" if existing.fired_at else ""
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Человек с таким ФИО уже числится за управлением «{owner}»{suffix}",
        )

    person = Person(
        full_name=      person_in.full_name,
        rank=           person_in.rank           or None,
        doc_number=     person_in.doc_number     or None,
        department=     department,
        position_title= person_in.position_title or None,
        birth_date=     person_in.birth_date,
        phone=          person_in.phone          or None,
        notes=          person_in.notes          or None,
    )
    db.add(person)
    try:
        db.flush()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=status.HTTP_409_CONFLICT,
                            detail="Человек с таким ФИО уже есть в базе")

    # Уведомляем админов — department добавил нового человека в общую базу.
    # Админ видит это событие сразу в колокольчике, без нужды искать вручную.
    admin_recipients: list[int] = []
    if current_user.role != "admin":
        admin_recipients = notify_all_admins(
            db,
            kind  = "person_applied",
            title = f"«{current_user.username}» добавил(а) человека в базу",
            body  = (f"ФИО: {person.full_name}"
                     + (f" · {person.rank}" if person.rank else "")),
            link  = None,
            exclude_user_id = current_user.id,
        )

    db.commit()
    db.refresh(person)

    for uid in admin_recipients:
        await manager.push_to_user(uid, {
            "action": "notification_new", "kind": "person_applied",
        })
    return person


@router.put("/{person_id}", response_model=PersonResponse, summary="Обновить данные человека")
async def update_person(
        person_id: int,
        person_in: PersonUpdate,
        db:        Session = Depends(get_db),
        current_user: User = Depends(require_permission("persons")),
):
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Запись не найдена")
    if current_user.role != "admin" and person.department != current_user.username:
        raise HTTPException(status_code=403, detail="Нет доступа")

    for field in ("full_name", "rank", "doc_number", "position_title",
                  "birth_date", "phone", "notes"):
        val = getattr(person_in, field, None)
        if val is not None:
            setattr(person, field, val or None)

    if person_in.department is not None and current_user.role == "admin":
        person.department = person_in.department or None

    db.commit()
    db.refresh(person)

    # Уведомляем подключённых клиентов: списки, в которых эта персона
    # является держателем (МНИ и др.) — должны обновиться online.
    await manager.broadcast({"action": "person_update", "person_id": person.id})

    return person


@router.delete("/{person_id}", summary="Удалить человека из базы (hard-delete)")
def delete_person(
        person_id: int,
        db:        Session = Depends(get_db),
        current_user: User = Depends(require_permission("persons")),
):
    """
    Физическое удаление. Стирает Person + каскадно duty_schedule_persons
    и duty_marks (ondelete=CASCADE), но не slots (там денормализованное ФИО
    остаётся). Используется для уборки ошибочно созданных записей.

    Для обычного «увольнения» (с сохранением истории) применяйте
    POST /persons/{id}/fire — он оставит Person и duty_marks, удалит только
    активные duty_schedule_persons.
    """
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Человек не найден")
    if current_user.role != "admin" and person.department != current_user.username:
        raise HTTPException(status_code=403, detail="Нет доступа")
    db.delete(person)
    db.commit()
    return {"message": "Удалён из базы"}


# ─── Увольнение и восстановление (admin-only) ────────────────────────────────

@router.post(
    "/{person_id}/fire",
    response_model=PersonResponse,
    summary="Уволить (admin-only)",
)
def fire_person(
        person_id:     int,
        db:            Session = Depends(get_db),
        current_admin: User    = Depends(get_current_active_admin),
):
    """
    Мягкое увольнение:
      • Person.fired_at = now()  — запись остаётся в базе.
      • Удаляются все duty_schedule_persons для этого человека
        (уволенный выпадает из активных графиков наряда).
      • duty_marks НЕ удаляются — история отметок сохраняется.
      • slots НЕ трогаем — там денормализованное ФИО.

    Идемпотентно: повторный вызов на уволенном возвращает 409.
    Централизованно у админа — департаменты увольнять не могут.
    """
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Человек не найден")
    if person.fired_at is not None:
        raise HTTPException(status_code=409, detail="Уже уволен")

    person.fired_at = datetime.now(timezone.utc)

    # Удаляем из всех активных графиков нарядов одним запросом.
    # duty_marks остаются — это история.
    db.query(DutySchedulePerson) \
      .filter(DutySchedulePerson.person_id == person_id) \
      .delete(synchronize_session=False)

    db.commit()
    db.refresh(person)
    return person


@router.post(
    "/{person_id}/unfire",
    response_model=PersonResponse,
    summary="Восстановить уволенного (admin-only)",
)
def unfire_person(
        person_id:     int,
        db:            Session = Depends(get_db),
        current_admin: User    = Depends(get_current_active_admin),
):
    """
    Отмена увольнения: fired_at → NULL. В активные графики не возвращаем
    автоматически — админ/управление заново добавит при необходимости.
    Идемпотентно: повторный вызов на активном вернёт 409.
    """
    person = db.query(Person).filter(Person.id == person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Человек не найден")
    if person.fired_at is None:
        raise HTTPException(status_code=409, detail="Не уволен")

    person.fired_at = None
    db.commit()
    db.refresh(person)
    return person


# ─── Внутренняя функция: upsert при заполнении слота ─────────────────────────

def upsert_person_from_slot(
        db:         Session,
        full_name:  str,
        rank:       str | None,
        doc_number: str | None,
        department: str | None = None,
) -> None:
    """
    Создаёт или обновляет запись в базе людей при сохранении слота.
    ИСПРАВЛЕНО: использует pg_insert с on_conflict_do_update
    вместо SELECT + условного INSERT (два круговых пути к БД → один).
    """
    if not full_name or not full_name.strip():
        return

    full_name = full_name.strip()
    now = datetime.now(timezone.utc)

    stmt = pg_insert(Person).values(
        full_name=  full_name,
        rank=       rank.strip()       if rank       else None,
        doc_number= doc_number.strip() if doc_number else None,
        department= department.strip() if department else None,
        created_at= now,
        updated_at= now,
    )
    # Правила обновления при конфликте по full_name:
    #   rank / doc_number — НЕ затираем уже заполненные значения
    #     (если в базе есть звание, а новый слот его не прислал — сохраняем).
    #   department        — НАОБОРОТ: всегда обновляем на присланное, если
    #     оно не null. То есть «человек сейчас принадлежит тому управлению,
    #     кто его последним заполнил в слоте». Это требование бизнес-логики:
    #     если upr_5 внёс ФИО — в базе он становится «upr_5», независимо
    #     от того, что там было раньше.
    stmt = stmt.on_conflict_do_update(
        index_elements=["full_name"],
        set_={
            "rank":       text("COALESCE(persons.rank,       EXCLUDED.rank)"),
            "doc_number": text("COALESCE(persons.doc_number, EXCLUDED.doc_number)"),
            "department": text("COALESCE(EXCLUDED.department, persons.department)"),
            "updated_at": stmt.excluded.updated_at,
        },
    )

    try:
        db.execute(stmt)
        db.flush()
    except Exception:
        db.rollback()


# ─── Расхождения данных (конфликты от тестирования и т.п.) ──────────────────
# Когда участник заполняет анкету в проф.подготовке, система пытается влить
# его данные в Person. Если значения отличаются от уже сохранённых —
# создаётся PersonDataConflict. Решает админ: оставить старое или применить
# новое. Записи с NULL resolved_at — открытые, ждут решения.

from typing import Literal as _Literal           # noqa: E402  (ниже)
from app.models.person_conflict import (         # noqa: E402
    PersonDataConflict, CONFLICT_FIELDS, CONFLICT_CHOICES,
)


CONFLICT_FIELD_LABELS = {
    "department":     "Управление / отдел",
    "phone":          "Телефон",
    "position_title": "Должность",
}


class ConflictOut(BaseModel):
    id:               int
    person_id:        int
    person_full_name: Optional[str] = None
    field_name:       str
    field_label:      str
    old_value:        Optional[str] = None
    new_value:        Optional[str] = None
    source:           str
    attempt_id:       Optional[int] = None
    created_at:       datetime
    resolved_at:      Optional[datetime] = None
    resolved_by:      Optional[str] = None
    resolved_choice:  Optional[str] = None

    model_config = ConfigDict(from_attributes=True)


class ConflictResolveIn(BaseModel):
    choice: _Literal["old", "new"]


@router.get("/conflicts", response_model=List[ConflictOut],
            summary="Список расхождений данных в общей базе людей (для админа)")
def list_person_conflicts(
        only_pending: bool = Query(True,
                description="Только нерешённые (resolved_at IS NULL)"),
        db:           Session = Depends(get_db),
        current_admin: User    = Depends(get_current_active_admin),
):
    q = db.query(PersonDataConflict)
    if only_pending:
        q = q.filter(PersonDataConflict.resolved_at.is_(None))
    items = q.order_by(PersonDataConflict.created_at.desc()).limit(500).all()
    out = []
    for c in items:
        out.append(ConflictOut(
            id=c.id,
            person_id=c.person_id,
            person_full_name=(c.person.full_name if c.person else None),
            field_name=c.field_name,
            field_label=CONFLICT_FIELD_LABELS.get(c.field_name, c.field_name),
            old_value=c.old_value,
            new_value=c.new_value,
            source=c.source,
            attempt_id=c.attempt_id,
            created_at=c.created_at,
            resolved_at=c.resolved_at,
            resolved_by=c.resolved_by,
            resolved_choice=c.resolved_choice,
        ))
    return out


@router.get("/conflicts/count",
            summary="Сколько открытых расхождений ждут решения")
def count_person_conflicts(
        db:            Session = Depends(get_db),
        current_admin: User    = Depends(get_current_active_admin),
):
    n = (db.query(PersonDataConflict)
           .filter(PersonDataConflict.resolved_at.is_(None))
           .count())
    return {"count": n}


@router.post("/conflicts/{conflict_id}/resolve", response_model=ConflictOut,
             summary="Разрешить расхождение: 'old' оставит старое, 'new' — применит новое")
async def resolve_person_conflict(
        conflict_id:   int,
        payload:       ConflictResolveIn,
        db:            Session = Depends(get_db),
        current_admin: User    = Depends(get_current_active_admin),
):
    c = db.query(PersonDataConflict).filter(PersonDataConflict.id == conflict_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Расхождение не найдено")
    if c.resolved_at:
        raise HTTPException(status_code=409, detail="Уже разрешено")
    if c.field_name not in CONFLICT_FIELDS:
        raise HTTPException(status_code=400,
                            detail=f"Неизвестное поле «{c.field_name}»")

    # Применяем решение
    person = db.query(Person).filter(Person.id == c.person_id).first()
    if not person:
        raise HTTPException(status_code=404, detail="Человек удалён из базы")

    if payload.choice == "new":
        setattr(person, c.field_name, c.new_value)

    c.resolved_at     = datetime.now(timezone.utc)
    c.resolved_by     = current_admin.username
    c.resolved_choice = payload.choice
    db.commit()
    db.refresh(c)

    # Уведомим всех клиентов о person_update — в учёте МНИ, заявках и пр.
    # денормализованные ФИО подтянутся автоматически.
    if payload.choice == "new":
        try:
            await manager.broadcast({
                "action": "person_update",
                "person_id": person.id,
            })
        except Exception:
            pass

    return ConflictOut(
        id=c.id,
        person_id=c.person_id,
        person_full_name=(c.person.full_name if c.person else None),
        field_name=c.field_name,
        field_label=CONFLICT_FIELD_LABELS.get(c.field_name, c.field_name),
        old_value=c.old_value,
        new_value=c.new_value,
        source=c.source,
        attempt_id=c.attempt_id,
        created_at=c.created_at,
        resolved_at=c.resolved_at,
        resolved_by=c.resolved_by,
        resolved_choice=c.resolved_choice,
    )
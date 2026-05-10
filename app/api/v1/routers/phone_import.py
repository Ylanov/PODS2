# app/api/v1/routers/phone_import.py
"""
Импорт номеров телефона из Excel-файла в Базу людей.

Воркфлоу:
  1. Админ выбирает .xlsx (структура: ФИО / Служебный / Домашний / Мобильный).
  2. POST /admin/persons/import-phones/preview — парсим, ищем Person по ФИО:
       • exact match (нормализованный full_name) → 'matched'
       • не нашли точно, но есть похожие → 'ambiguous' (несколько кандидатов)
       • вообще не нашли → 'unknown'
     Возвращаем структуру для UI.
  3. Админ в диалоге:
       • для matched — отмечает галки (по умолчанию ✓ если phone пуст)
       • для ambiguous — выбирает radio-кандидата
       • unknown — пропускает (не создаём людей автоматически)
  4. POST /admin/persons/import-phones/apply — применяем изменения.

В Person есть только одно поле phone (а в Excel — три). Берём первый
непустой по приоритету: Мобильный > Служебный > Домашний.
"""

import io
import re
import logging
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Request, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from app.api.dependencies import get_current_active_admin
from app.core.config import settings
from app.core.limiter import limiter
from app.core.websockets import manager
from app.db.database import get_db
from app.models.person import Person
from app.models.user import User


logger = logging.getLogger(__name__)
router = APIRouter()


# ─── Pydantic ────────────────────────────────────────────────────────────────

class _PersonRef(BaseModel):
    id:        int
    full_name: str
    rank:      Optional[str] = None
    department: Optional[str] = None
    phone:     Optional[str] = None


class MatchedRow(BaseModel):
    """ФИО точно совпало с одним Person."""
    excel_name:   str
    person:       _PersonRef
    new_phone:    str
    has_old_phone: bool = False


class AmbiguousRow(BaseModel):
    """ФИО неоднозначно — несколько кандидатов, ждём ручного выбора."""
    excel_name:   str
    new_phone:    str
    candidates:   list[_PersonRef]


class UnknownRow(BaseModel):
    """ФИО вообще не нашли в Базе людей."""
    excel_name: str
    new_phone:  str


class PreviewResponse(BaseModel):
    matched:    list[MatchedRow]
    ambiguous:  list[AmbiguousRow]
    unknown:    list[UnknownRow]
    total_rows: int
    skipped_no_phone: int = 0
    skipped_no_name:  int = 0


class ApplyChange(BaseModel):
    person_id: int
    phone:     str = Field(..., min_length=1, max_length=50)


class ApplyPayload(BaseModel):
    changes: list[ApplyChange]


# ─── Helpers ─────────────────────────────────────────────────────────────────

_NAME_NORMALIZE_RE = re.compile(r"\s+")


def _normalize_name(s: str) -> str:
    """Нижний регистр + схлопнутые пробелы для сравнения ФИО."""
    return _NAME_NORMALIZE_RE.sub(" ", (s or "").strip()).lower()


def _normalize_one_phone(raw: str) -> str:
    """
    Нормализация ОДНОГО номера: оставляем только цифры, приводим к +7XXXXXXXXXX.
    Excel-формат «89260367265» (числом) → «+79260367265».
    Если меньше 10 цифр — отдаём как есть (мог быть короткий служебный).
    """
    s = (raw or "").strip()
    if not s:
        return ""
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if len(digits) == 11 and digits[0] in ("7", "8"):
        return "+7" + digits[1:]
    if len(digits) == 10:
        return "+7" + digits
    return s


# Разделители номеров когда в одной ячейке их несколько: запятая, точка с
# запятой, перенос строки, слэш, «или», два и более пробела.
_PHONE_SPLIT_RE = re.compile(r"[,\n;/]+|\s+или\s+|\s{2,}")


def _collect_phones(*cells) -> str:
    """
    Собирает все номера из всех ячеек строки (служебный/домашний/мобильный)
    в одну строку через «, ». Если в ячейке несколько номеров — разбивает.
    Дубликаты отбрасываются. По задумке: у Person.phone одна строка, в
    которой ВСЕ известные номера человека (мобильный, служебный, домашний —
    различия не нужны, главное что номер этого человека).
    """
    seen: set[str] = set()
    out:  list[str] = []
    for cell in cells:
        if cell is None:
            continue
        # Excel может прислать число (89260367265) или текст
        raw = str(cell).strip()
        if not raw:
            continue
        # Если в ячейке несколько номеров — разбиваем
        parts = _PHONE_SPLIT_RE.split(raw)
        for part in parts:
            norm = _normalize_one_phone(part)
            if norm and norm not in seen:
                seen.add(norm)
                out.append(norm)
    return ", ".join(out)


def _person_ref(p: Person) -> _PersonRef:
    return _PersonRef(
        id=p.id, full_name=p.full_name, rank=p.rank,
        department=p.department, phone=p.phone,
    )


def _find_candidates(excel_name: str, persons_index: dict) -> list[Person]:
    """
    Возвращает кандидатов:
      • exact (по нормализованному имени) → [single]
      • иначе ищем по «фамилия имя» (первые два слова) — все совпадения
    persons_index: {normalized_full_name: [Person]} (несколько на случай
    тёзок).
    """
    norm = _normalize_name(excel_name)
    if not norm:
        return []
    # Точное совпадение
    if norm in persons_index:
        return persons_index[norm]
    # По фамилии и имени (первые два слова) — может быть совпадение
    # «Иванов Иван» с базой «Иванов Иван Иванович». Используем prefix-match.
    parts = norm.split()
    if len(parts) >= 2:
        prefix = " ".join(parts[:2])
        out = []
        for full_norm, ps in persons_index.items():
            if full_norm.startswith(prefix + " ") or full_norm == prefix:
                out.extend(ps)
        if out:
            return out
    return []


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.post("/persons/import-phones/preview", response_model=PreviewResponse,
             summary="Парсинг Excel с телефонами + автоматический матчинг ФИО")
@limiter.limit(lambda: settings.IMPORT_RATE_LIMIT)
async def preview(
    request: Request,
    file:  UploadFile = File(...),
    db:    Session = Depends(get_db),
    admin: User    = Depends(get_current_active_admin),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Ожидается .xlsx файл")

    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать Excel: {exc}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    # Пропускаем шапку — первая строка с заголовками.
    try:
        header = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=400, detail="Пустой файл")

    # Индекс активных людей по нормализованному ФИО.
    all_persons = (
        db.query(Person)
        .filter(Person.fired_at.is_(None))
        .all()
    )
    persons_index: dict[str, list[Person]] = {}
    for p in all_persons:
        persons_index.setdefault(_normalize_name(p.full_name), []).append(p)

    matched:    list[MatchedRow] = []
    ambiguous:  list[AmbiguousRow] = []
    unknown:    list[UnknownRow] = []
    skipped_no_phone = 0
    skipped_no_name  = 0
    total_rows       = 0

    for row in rows_iter:
        if not row or len(row) < 2:
            continue
        # row = (ФИО, служебный, домашний, мобильный)
        name = (row[0] or "").strip() if isinstance(row[0], str) else (str(row[0]).strip() if row[0] else "")
        if not name:
            skipped_no_name += 1
            continue
        # Все колонки — это просто разные номера одного человека.
        # Пользователь явно сказал: домашний/мобильный/служебный — это
        # всё «телефоны» без разделения, складываем все в одно поле.
        phone = _collect_phones(*row[1:])
        if not phone:
            skipped_no_phone += 1
            continue
        total_rows += 1

        candidates = _find_candidates(name, persons_index)
        if len(candidates) == 1:
            p = candidates[0]
            matched.append(MatchedRow(
                excel_name=name,
                person=_person_ref(p),
                new_phone=phone,
                has_old_phone=bool((p.phone or "").strip()),
            ))
        elif len(candidates) > 1:
            ambiguous.append(AmbiguousRow(
                excel_name=name,
                new_phone=phone,
                candidates=[_person_ref(p) for p in candidates],
            ))
        else:
            unknown.append(UnknownRow(excel_name=name, new_phone=phone))

    return PreviewResponse(
        matched=matched,
        ambiguous=ambiguous,
        unknown=unknown,
        total_rows=total_rows,
        skipped_no_phone=skipped_no_phone,
        skipped_no_name=skipped_no_name,
    )


@router.post("/persons/import-phones/apply",
             summary="Применить телефоны: записывает Person.phone у выбранных")
@limiter.limit(lambda: settings.IMPORT_RATE_LIMIT)
async def apply(
    request: Request,
    payload: ApplyPayload,
    db:      Session = Depends(get_db),
    admin:   User    = Depends(get_current_active_admin),
):
    if not payload.changes:
        return {"updated": 0}

    person_ids = [c.person_id for c in payload.changes]
    rows = db.query(Person).filter(Person.id.in_(person_ids)).all()
    by_id = {p.id: p for p in rows}

    updated_ids: list[int] = []
    for ch in payload.changes:
        p = by_id.get(ch.person_id)
        if not p:
            continue
        new_phone = ch.phone.strip()
        if p.phone == new_phone:
            continue
        p.phone = new_phone
        updated_ids.append(p.id)
    db.commit()

    for pid in updated_ids:
        await manager.broadcast({"action": "person_update", "person_id": pid})

    return {"updated": len(updated_ids), "received": len(payload.changes)}
